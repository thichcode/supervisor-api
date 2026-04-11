"""
File Processor - Read and extract data from various file formats
Supports: PDF, Excel (xlsx/xls), CSV, JSON, Text, Images (OCR)
"""

import json
import csv
import io
import re
import hashlib
from pathlib import Path
from typing import Optional, List, Dict, Any, Union
from dataclasses import dataclass, field
from datetime import datetime
import structlog

logger = structlog.get_logger()

# Try importing libraries
PANDAS_AVAILABLE = False
OPENPYXL_AVAILABLE = False
PDFPLUMBER_AVAILABLE = False
PYPDF_AVAILABLE = False
PIL_AVAILABLE = False
PYTESSERACT_AVAILABLE = False
PADDLEOCR_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    pass

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    pass

try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    pass

try:
    from pypdf import PdfReader
    PYPDF_AVAILABLE = True
except ImportError:
    pass

try:
    from PIL import Image
    PIL_AVAILABLE = True
except ImportError:
    pass

try:
    import pytesseract
    PYTESSERACT_AVAILABLE = True
except ImportError:
    pass

try:
    from paddleocr import PaddleOCR
    PADDLEOCR_AVAILABLE = True
except ImportError:
    pass


@dataclass
class FileContent:
    """Extracted file content"""
    filename: str
    content: str
    content_type: str  # "text", "table", "structured", "image"
    metadata: Dict[str, Any] = field(default_factory=dict)
    extracted_at: datetime = field(default_factory=datetime.utcnow)
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            "filename": self.filename,
            "content": self.content,
            "content_type": self.content_type,
            "metadata": self.metadata,
            "extracted_at": self.extracted_at.isoformat(),
        }


@dataclass
class TableData:
    """Table data extracted from files"""
    headers: List[str]
    rows: List[List[Any]]
    sheet_name: Optional[str] = None
    
    def to_dataframe(self):
        """Convert to pandas DataFrame"""
        if PANDAS_AVAILABLE:
            return pd.DataFrame(self.rows, columns=self.headers)
        return None
    
    def to_markdown(self) -> str:
        """Convert to markdown table"""
        if not self.rows:
            return ""
        
        lines = []
        lines.append("| " + " | ".join(self.headers) + " |")
        lines.append("| " + " | ".join(["---"] * len(self.headers)) + " |")
        
        for row in self.rows:
            values = [str(v) if v is not None else "" for v in row]
            lines.append("| " + " | ".join(values) + " |")
        
        return "\n".join(lines)


class FileProcessor:
    """
    Multi-format file processor with OCR support
    """
    
    # Supported file types
    SUPPORTED_TYPES = {
        # Documents
        ".pdf": "pdf",
        ".txt": "text",
        ".md": "text",
        ".rst": "text",
        # Spreadsheets
        ".xlsx": "excel",
        ".xls": "excel",
        ".csv": "csv",
        ".tsv": "csv",
        # Structured
        ".json": "json",
        ".xml": "xml",
        ".yaml": "yaml",
        ".yml": "yaml",
        # Images
        ".png": "image",
        ".jpg": "image",
        ".jpeg": "image",
        ".tiff": "image",
        ".tif": "image",
        ".bmp": "image",
        ".gif": "image",
    }
    
    def __init__(
        self,
        ocr_enabled: bool = True,
        ocr_language: str = "eng+vie",  # English + Vietnamese
        max_file_size: int = 50 * 1024 * 1024,  # 50MB
        extract_tables: bool = True,
    ):
        self.ocr_enabled = ocr_enabled
        self.ocr_language = ocr_language
        self.max_file_size = max_file_size
        self.extract_tables = extract_tables
        
        # Initialize OCR
        self._ocr_engine = None
    
    def _get_ocr_engine(self):
        """Lazy load OCR engine"""
        if self._ocr_engine is None:
            if PADDLEOCR_AVAILABLE:
                self._ocr_engine = "paddle"
                logger.info("Using PaddleOCR engine")
            elif PYTESSERACT_AVAILABLE:
                self._ocr_engine = "tesseract"
                logger.info("Using Tesseract OCR engine")
            else:
                self._ocr_engine = "none"
                logger.warning("No OCR engine available")
        
        return self._ocr_engine
    
    def process_file(self, file_path: Union[str, Path]) -> FileContent:
        """
        Process a file and extract content
        
        Args:
            file_path: Path to the file
            
        Returns:
            FileContent with extracted data
        """
        path = Path(file_path)
        
        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")
        
        if path.stat().st_size > self.max_file_size:
            raise ValueError(f"File too large: {path.stat().st_size} bytes (max: {self.max_file_size})")
        
        suffix = path.suffix.lower()
        
        if suffix not in self.SUPPORTED_TYPES:
            raise ValueError(f"Unsupported file type: {suffix}")
        
        file_type = self.SUPPORTED_TYPES[suffix]
        
        # Process based on type
        if file_type == "pdf":
            return self._process_pdf(path)
        elif file_type == "excel":
            return self._process_excel(path)
        elif file_type == "csv":
            return self._process_csv(path)
        elif file_type == "json":
            return self._process_json(path)
        elif file_type in ("text", "md", "rst"):
            return self._process_text(path)
        elif file_type == "image":
            return self._process_image(path)
        else:
            return self._process_text(path)
    
    def _process_pdf(self, path: Path) -> FileContent:
        """Extract text and tables from PDF"""
        content_parts = []
        metadata = {"pages": 0, "tables": 0}
        
        if PDFPLUMBER_AVAILABLE:
            with pdfplumber.open(path) as pdf:
                metadata["pages"] = len(pdf.pages)
                
                for page_num, page in enumerate(pdf.pages):
                    # Extract text
                    text = page.extract_text()
                    if text:
                        content_parts.append(f"--- Page {page_num + 1} ---\n{text}")
                    
                    # Extract tables
                    if self.extract_tables:
                        tables = page.extract_tables()
                        for table_num, table in enumerate(tables):
                            if table:
                                metadata["tables"] += 1
                                table_obj = TableData(
                                    headers=table[0] if table else [],
                                    rows=table[1:] if len(table) > 1 else [],
                                    sheet_name=f"Page {page_num + 1}, Table {table_num + 1}"
                                )
                                content_parts.append(f"\n[Table from Page {page_num + 1}]:\n{table_obj.to_markdown()}")
        
        elif PYPDF_AVAILABLE:
            reader = PdfReader(path)
            metadata["pages"] = len(reader.pages)
            
            for page_num, page in enumerate(reader.pages):
                text = page.extract_text()
                if text:
                    content_parts.append(f"--- Page {page_num + 1} ---\n{text}")
        
        content = "\n\n".join(content_parts)
        
        return FileContent(
            filename=path.name,
            content=content,
            content_type="table" if metadata["tables"] > 0 else "text",
            metadata=metadata
        )
    
    def _process_excel(self, path: Path) -> FileContent:
        """Extract data from Excel files"""
        all_tables = []
        metadata = {"sheets": 0}
        
        if PANDAS_AVAILABLE:
            # Read all sheets
            excel_file = pd.ExcelFile(path)
            metadata["sheets"] = len(excel_file.sheet_names)
            metadata["sheet_names"] = excel_file.sheet_names
            
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(path, sheet_name=sheet_name)
                
                table = TableData(
                    headers=df.columns.tolist(),
                    rows=df.values.tolist(),
                    sheet_name=sheet_name
                )
                all_tables.append(table)
        
        elif OPENPYXL_AVAILABLE:
            wb = openpyxl.load_workbook(path)
            metadata["sheets"] = len(wb.sheetnames)
            metadata["sheet_names"] = wb.sheetnames
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Get headers
                headers = [cell.value for cell in ws[1]]
                
                # Get rows
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(list(row))
                
                table = TableData(
                    headers=headers,
                    rows=rows,
                    sheet_name=sheet_name
                )
                all_tables.append(table)
        
        # Format content
        content_parts = []
        for table in all_tables:
            content_parts.append(f"=== Sheet: {table.sheet_name} ===")
            content_parts.append(table.to_markdown())
            content_parts.append("")
        
        return FileContent(
            filename=path.name,
            content="\n".join(content_parts),
            content_type="table",
            metadata=metadata
        )
    
    def _process_csv(self, path: Path) -> FileContent:
        """Extract data from CSV/TSV files"""
        metadata = {"rows": 0, "columns": 0}
        
        if PANDAS_AVAILABLE:
            # Auto-detect delimiter
            df = pd.read_csv(path)
            
            metadata["rows"] = len(df)
            metadata["columns"] = len(df.columns)
            metadata["column_names"] = df.columns.tolist()
            
            table = TableData(
                headers=df.columns.tolist(),
                rows=df.values.tolist()
            )
            content = table.to_markdown()
        else:
            # Fallback to standard csv module
            with open(path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                rows = list(reader)
                
                if rows:
                    headers = rows[0]
                    data_rows = rows[1:]
                    
                    metadata["rows"] = len(data_rows)
                    metadata["columns"] = len(headers)
                    
                    table = TableData(headers=headers, rows=data_rows)
                    content = table.to_markdown()
                else:
                    content = ""
        
        return FileContent(
            filename=path.name,
            content=content,
            content_type="table",
            metadata=metadata
        )
    
    def _process_json(self, path: Path) -> FileContent:
        """Extract data from JSON files"""
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        metadata = {"type": type(data).__name__}
        
        if isinstance(data, dict):
            metadata["keys"] = list(data.keys())
            content = json.dumps(data, indent=2, ensure_ascii=False, default=str)
        elif isinstance(data, list):
            metadata["length"] = len(data)
            content = json.dumps(data, indent=2, ensure_ascii=False, default=str)
        else:
            content = str(data)
        
        return FileContent(
            filename=path.name,
            content=content,
            content_type="structured",
            metadata=metadata
        )
    
    def _process_text(self, path: Path) -> FileContent:
        """Extract text from plain text files"""
        with open(path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        metadata = {
            "size": path.stat().st_size,
            "lines": content.count('\n') + 1,
            "words": len(content.split()),
        }
        
        return FileContent(
            filename=path.name,
            content=content,
            content_type="text",
            metadata=metadata
        )
    
    def _process_image(self, path: Path) -> FileContent:
        """Extract text from images using OCR"""
        if not self.ocr_enabled:
            return FileContent(
                filename=path.name,
                content="[OCR disabled]",
                content_type="image",
                metadata={"error": "OCR disabled"}
            )
        
        ocr_engine = self._get_ocr_engine()
        
        if ocr_engine == "none":
            return FileContent(
                filename=path.name,
                content="[No OCR engine available]",
                content_type="image",
                metadata={"error": "No OCR engine installed"}
            )
        
        content = ""
        metadata = {"engine": ocr_engine}
        
        if ocr_engine == "paddle" and PADDLEOCR_AVAILABLE:
            try:
                if self._ocr_engine and hasattr(self._ocr_engine, 'ocr'):
                    pass  # Already initialized
                else:
                    self._paddle_ocr = PaddleOCR(use_angle_cls=True, lang='en', use_gpu=False)
                
                result = self._paddle_ocr.ocr(str(path))
                
                if result:
                    for line in result[0]:
                        content += line[1][0] + "\n"
                
                metadata["lines"] = len(result[0]) if result else 0
            except Exception as e:
                logger.error("PaddleOCR failed", error=str(e))
                content = f"[OCR Error: {str(e)}]"
        
        elif ocr_engine == "tesseract" and PYTESSERACT_AVAILABLE and PIL_AVAILABLE:
            try:
                img = Image.open(path)
                content = pytesseract.image_to_string(img, lang=self.ocr_language)
                metadata["size"] = img.size
            except Exception as e:
                logger.error("Tesseract OCR failed", error=str(e))
                content = f"[OCR Error: {str(e)}]"
        
        return FileContent(
            filename=path.name,
            content=content.strip(),
            content_type="text",
            metadata=metadata
        )
    
    def process_multiple(self, file_paths: List[Union[str, Path]]) -> List[FileContent]:
        """Process multiple files"""
        results = []
        for path in file_paths:
            try:
                result = self.process_file(path)
                results.append(result)
            except Exception as e:
                logger.error("File processing failed", file=str(path), error=str(e))
                results.append(FileContent(
                    filename=str(path),
                    content=f"[Error: {str(e)}]",
                    content_type="error",
                    metadata={"error": str(e)}
                ))
        return results
    
    def extract_structured_data(self, file_path: Union[str, Path]) -> Dict[str, Any]:
        """
        Extract structured data from file (for IT reports)
        
        Returns dict with common IT report fields:
        - headers
        - rows
        - summary (if available)
        """
        content = self.process_file(file_path)
        
        result = {
            "filename": content.filename,
            "content_type": content.content_type,
            "metadata": content.metadata,
            "text": content.content,
        }
        
        # Parse tables if available
        if content.content_type == "table":
            lines = content.content.split("\n")
            
            # Find markdown table
            table_start = -1
            for i, line in enumerate(lines):
                if line.startswith("|"):
                    table_start = i
                    break
            
            if table_start >= 0:
                # Parse markdown table
                table_lines = []
                for i in range(table_start, len(lines)):
                    if lines[i].startswith("|"):
                        table_lines.append(lines[i])
                    elif table_lines and lines[i].strip():
                        break
                
                # Extract headers and rows
                headers = []
                rows = []
                
                if len(table_lines) >= 2:
                    headers = [h.strip() for h in table_lines[0].split("|")[1:-1]]
                    
                    for line in table_lines[2:]:  # Skip separator
                        cells = [c.strip() for c in line.split("|")[1:-1]]
                        if cells:
                            rows.append(cells)
                
                result["table"] = {
                    "headers": headers,
                    "rows": rows,
                    "row_count": len(rows),
                }
        
        return result


# Global processor instance
_file_processor: Optional[FileProcessor] = None


def get_file_processor(**kwargs) -> FileProcessor:
    """Get or create global file processor"""
    global _file_processor
    if _file_processor is None:
        _file_processor = FileProcessor(**kwargs)
    return _file_processor
